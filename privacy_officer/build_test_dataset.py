import csv
from collections import Counter

rows = [
    # theme, te_detecteren (pipe-sep literal values), pii_type (pipe-sep), feedback_text
    # --- Slow grading by Smith ---
    ("Slow grading by Smith", "Smith", "naam",
     "Docent Smith heeft mijn opdracht al drie weken niet nagekeken. Dit is echt frustrerend voor de hele klas."),
    ("Slow grading by Smith", "Smith", "naam",
     "Meneer Smith geeft nooit op tijd feedback. Mijn inlevering van vorige maand heeft nog steeds geen beoordeling."),
    ("Slow grading by Smith", "Smith", "naam",
     "Smith corrects assignments way too slowly. We have been waiting over four weeks for our grades."),
    ("Slow grading by Smith", "Smith|j.smith@fontys.nl", "naam|email",
     "Ik heb Smith gemaild op j.smith@fontys.nl maar geen reactie gekregen over mijn uitstaande cijfer."),
    ("Slow grading by Smith", "Smith|smith@fontys.nl", "naam|email",
     "Our group emailed smith@fontys.nl twice about the missing grade but received no response from Smith."),
    ("Slow grading by Smith", "Smith|547123", "naam|studentnummer",
     "Student 547123 hier: de beoordeling van Smith duurt elke keer veel te lang, dit blokkeert mijn voortgang."),
    ("Slow grading by Smith", "Smith|1234567", "naam|studentnummer",
     "Studentnummer 1234567, de feedback van docent Smith op mijn portfolio is al vijf weken uit."),
    ("Slow grading by Smith", "Smith|lokaal 3.05", "naam|gebouw",
     "We moesten naar lokaal 3.05 lopen om Smith persoonlijk te vragen naar onze cijfers want hij reageert niet online."),
    ("Slow grading by Smith", "kale docent|derde etage", "indirect_fysiek|indirect_locatie",
     "De kale docent op de derde etage neemt weken om opdrachten na te kijken en communiceert hier niet over."),
    ("Slow grading by Smith", "lange man met bril", "indirect_fysiek",
     "De lange man met bril die Software Engineering geeft corrigeert structureel te laat en geeft geen uitleg."),

    # --- Insufficient contact hours ---
    ("Insufficient contact hours", "R10", "gebouw",
     "We hebben te weinig contactmomenten. De lessen vinden slechts twee keer per week plaats in R10."),
    ("Insufficient contact hours", "lokaal 2.05", "gebouw",
     "In lokaal 2.05 hebben we maar een contactuur per week, dat is echt onvoldoende voor zo een zwaar vak."),
    ("Insufficient contact hours", "TQ 3.14", "gebouw",
     "De contacturen in TQ 3.14 zijn minimaal. We kunnen onze vragen tijdens de les nauwelijks kwijt."),
    ("Insufficient contact hours", "R1", "gebouw",
     "The classes in R1 are only scheduled once a week. We really need more contact hours with the teacher."),
    ("Insufficient contact hours", "Jansen|gebouw R2", "naam|gebouw",
     "In gebouw R2 geeft mevrouw Jansen les, maar er zijn te weinig geplande contactmomenten per blok."),
    ("Insufficient contact hours", "lokaal 4.12", "gebouw",
     "We zitten altijd in lokaal 4.12 maar de lessen zijn te kort en de frequentie is veel te laag."),
    ("Insufficient contact hours", "Van den Berg|3e etage", "naam|indirect_locatie",
     "Docent Van den Berg op de 3e etage is zelden beschikbaar voor extra contacturen buiten de geplande lessen."),
    ("Insufficient contact hours", "2e etage", "indirect_locatie",
     "De docenten op de 2e etage zijn vaak afwezig, waardoor geplande contactmomenten regelmatig worden afgezegd."),
    ("Insufficient contact hours", "Peters", "naam",
     "Peters geeft maar drie uur contacttijd per week. Voor een vak van vijf ECTS is dat absoluut onvoldoende."),
    ("Insufficient contact hours", "Peters|rolstoelgebruiker|R10", "naam|toegankelijkheid|gebouw",
     "Als rolstoelgebruiker is lokaal R10 voor mij onbereikbaar vanwege de defecte lift. Docent Peters biedt geen alternatief contactmoment aan."),

    # --- Course depth & quality (positive) ---
    ("Course depth & quality (positive)", "De Vries", "naam",
     "Docent De Vries geeft heel diepgaande lessen. Je leert echt iets nieuws bij elke sessie, geweldig vak."),
    ("Course depth & quality (positive)", "Bakker", "naam",
     "Meneer Bakker legt de stof altijd helder uit en gaat diep genoeg in op de materie. Aanrader voor iedereen."),
    ("Course depth & quality (positive)", "Williams", "naam",
     "This course taught by professor Williams is excellent. The depth and quality of the content are outstanding."),
    ("Course depth & quality (positive)", "Smit", "naam",
     "Hartelijk dank aan mevrouw Smit voor de geweldige lessen. Dit was veruit het beste vak van mijn studiejaar."),
    ("Course depth & quality (positive)", "Visser|876543", "naam|studentnummer",
     "Student 876543: de kwaliteit van de lessen van docent Visser is uitstekend, ik heb enorm veel geleerd."),
    ("Course depth & quality (positive)", "Chen|612345@student.fontys.nl", "naam|email",
     "Als student (612345@student.fontys.nl) heb ik Chen persoonlijk bedankt. De manier waarop hij het materiaal presenteert is boeiend en goed doordacht."),
    ("Course depth & quality (positive)", "Mulder", "naam",
     "Docent Mulder slaagt erin complexe concepten begrijpelijk te maken zonder diepgang te verliezen. Top docent."),
    ("Course depth & quality (positive)", "Patel|556677", "naam|studentnummer",
     "Student 556677 hier: mevrouw Patel is een uitstekende docent. Haar colleges zijn altijd goed voorbereid en zeer leerzaam."),
    ("Course depth & quality (positive)", "korte blonde haren", "indirect_fysiek",
     "De vrouwelijke docent met korte blonde haren geeft verreweg de beste lessen van het blok dit semester."),
    ("Course depth & quality (positive)", "Smit|ADHD", "naam|gezondheid",
     "Docent Smit is de enige die echt rekening houdt met studenten met ADHD. Ze geeft altijd een heldere structuur en stuurt het materiaal van tevoren op."),

    # --- Unclear assessment criteria ---
    ("Unclear assessment criteria", "Janssen", "naam",
     "De beoordelingscriteria van docent Janssen zijn onduidelijk. We weten niet wat er precies verwacht wordt."),
    ("Unclear assessment criteria", "Peters", "naam",
     "Peters legt nooit goed uit hoe het cijfer is opgebouwd. Meer transparantie over de weging is nodig."),
    ("Unclear assessment criteria", "Johnson", "naam",
     "It is unclear what Johnson expects from us. The assessment criteria seem to change every single week."),
    ("Unclear assessment criteria", "Bosman", "naam",
     "Mevrouw Bosman hanteert onduidelijke criteria. De ene keer telt presentatie mee, de andere keer niet."),
    ("Unclear assessment criteria", "Vermeer|gebouw R10", "naam|gebouw",
     "In gebouw R10 heeft docent Vermeer ons nooit een duidelijke uitleg gegeven over de weging van opdrachten."),
    ("Unclear assessment criteria", "Willems|234567", "naam|studentnummer",
     "Docent Willems en ik (studentnummer 234567) hebben een gesprek gehad maar de criteria blijven vaag."),
    ("Unclear assessment criteria", "345678", "studentnummer",
     "Studentnummer 345678: ik begrijp de rubric niet. De criteria zijn tegenstrijdig en inconsistent opgesteld."),
    ("Unclear assessment criteria", "765432", "studentnummer",
     "Als student 765432 heb ik meerdere keren gevraagd om verduidelijking van de beoordelingscriteria, zonder succes."),
    ("Unclear assessment criteria", "m.de_jong@student.fontys.nl", "email",
     "Ik heb via m.de_jong@student.fontys.nl ook mijn medestudenten gepolst en niemand begrijpt de criteria."),
    ("Unclear assessment criteria", "Janssen|dyslexie", "naam|gezondheid",
     "Ik heb dyslexie en heb docent Janssen meerdere keren gevraagd om de rubric in een groter lettertype aan te leveren. Er is nooit op gereageerd."),

    # --- Poor teacher communication ---
    ("Poor teacher communication", "Hendriks|hendriks@fontys.nl", "naam|email",
     "Docent Hendriks is via hendriks@fontys.nl al twee weken niet bereikbaar voor een dringende vraag over mijn toets."),
    ("Poor teacher communication", "Claes|c.claes@fontys.nl", "naam|email",
     "Mevrouw Claes (c.claes@fontys.nl) antwoordt zelden binnen een week op vragen die via de mail worden gesteld."),
    ("Poor teacher communication", "Ooms|f.ooms@fontys.nl", "naam|email",
     "Meneer Ooms (f.ooms@fontys.nl) reageert zelden binnen een week en communiceert onduidelijk over deadlines."),
    ("Poor teacher communication", "Kuipers|06-87654321", "naam|telefoon",
     "De heer Kuipers (06-87654321) is telefonisch onbereikbaar en reageert ook niet op ingesproken voicemails."),
    ("Poor teacher communication", "Van Dam|06-12345678", "naam|telefoon",
     "Ik heb mevrouw Van Dam gebeld op 06-12345678 maar ze neemt nooit op en belt ook nooit terug."),
    ("Poor teacher communication", "Peters", "naam",
     "Peters communicates poorly. Announcements are always last-minute and the information is often confusing."),
    ("Poor teacher communication", "Lammers", "naam",
     "De communicatie van docent Lammers laat te wensen over. Roosterwijzigingen worden nooit op tijd doorgegeven."),
    ("Poor teacher communication", "Adriaans", "naam",
     "Mevrouw Adriaans communiceert inconsistent: soms via Teams, soms via mail, nooit op een vaste manier."),
    ("Poor teacher communication", "Van Dam|@v.vandam", "naam|username",
     "Docent Van Dam reageert niet op Teams (@v.vandam) en ook niet op zijn officiele e-mailadres."),
    ("Poor teacher communication", "06-23456789|depressie", "telefoon|gezondheid",
     "Ik heb de docentenlijn gebeld op 06-23456789 om aan te geven dat ik een depressie heb en extra begeleiding nodig heb. Er is nooit teruggebeld."),

    # --- Teaching style & explanation quality ---
    ("Teaching style & explanation quality", "Willems", "naam",
     "Docent Willems legt moeilijke stof altijd stap voor stap uit. Zijn manier van uitleggen is werkelijk top."),
    ("Teaching style & explanation quality", "Groot", "naam",
     "Mevrouw Groot gebruikt veel voorbeelden uit de praktijk, waardoor de theorie echt blijft hangen bij studenten."),
    ("Teaching style & explanation quality", "Adams", "naam",
     "Professor Adams explains concepts clearly and consistently uses real-world examples to make theory practical."),
    ("Teaching style & explanation quality", "Van Leeuwen", "naam",
     "De didactiek van mevrouw Van Leeuwen is uitstekend. Ze motiveert studenten om zelf kritisch na te denken."),
    ("Teaching style & explanation quality", "Roos", "naam",
     "De uitleg van docent Roos is soms onduidelijk. Hij gaat te snel door de stof heen en pauzeert niet voor vragen."),
    ("Teaching style & explanation quality", "Koster", "naam",
     "Docent Koster gaat ervan uit dat iedereen de basisstof al kent, maar dat klopt niet voor nieuwe studenten."),
    ("Teaching style & explanation quality", "Prins|TQ 2.08", "naam|gebouw",
     "In lokaal TQ 2.08 geeft meneer Prins les op een manier die voor de meeste studenten onbegrijpelijk is."),
    ("Teaching style & explanation quality", "Timmermans|R1", "naam|gebouw",
     "Mevrouw Timmermans in R1 legt het altijd te snel uit. Er is geen ruimte voor vragen of herhaling."),
    ("Teaching style & explanation quality", "Van Leeuwen|autisme", "naam|gezondheid",
     "Als student met autisme waardeer ik de vaste lesstructuur van mevrouw Van Leeuwen enorm. Ze kondigt altijd vooraf aan wat er gaat gebeuren."),
    ("Teaching style & explanation quality", "visuele beperking", "toegankelijkheid",
     "De presentaties zijn nooit toegankelijk voor studenten met een visuele beperking. Kleine letters, laag contrast en geen alternatieve tekst bij afbeeldingen."),

    # --- Workload & time management ---
    ("Workload & time management", "Van den Berg", "naam",
     "Docent Van den Berg geeft elke week drie grote opdrachten op. De werkdruk is voor geen enkele student te doen."),
    ("Workload & time management", "Bos", "naam",
     "Mevrouw Bos plant veel te veel in per blok. Studenten raken hierdoor structureel overbelast en gestrest."),
    ("Workload & time management", "Thompson", "naam",
     "The workload assigned by Thompson is completely unreasonable. Four major assignments in a single week is too much."),
    ("Workload & time management", "Garcia", "naam",
     "The time management of professor Garcia could be greatly improved. Assignments are always announced last-minute."),
    ("Workload & time management", "Smeets", "naam",
     "Docent Smeets houdt geen rekening met de werkdruk van andere vakken bij het plannen van zijn opdrachten."),
    ("Workload & time management", "Pieters|456789", "naam|studentnummer",
     "Ik (studentnummer 456789) heb docent Pieters gevraagd om spreiding van deadlines maar dit wordt steeds genegeerd."),
    ("Workload & time management", "Kuiper|654321", "naam|studentnummer",
     "Studentnummer 654321 wil aangeven dat docent Kuiper te veel huiswerk opgeeft voor een halfjaarsvak."),
    ("Workload & time management", "123456", "studentnummer",
     "Student 123456 hier. De deadlines van dit vak overlappen altijd met andere vakken, dit is niet werkbaar."),
    ("Workload & time management", "Franssen|gebouw R2|lokaal 1.04", "naam|gebouw|gebouw",
     "In gebouw R2 lokaal 1.04 geeft meneer Franssen een vak met een volledig onrealistische tijdsplanning."),
    ("Workload & time management", "Brouwer|burnout", "naam|gezondheid",
     "Docent Brouwer heeft geen begrip voor studenten die een burnout hebben of op de rand zitten. De werkdruk van dit vak is onverantwoord hoog."),

    # --- Student support & teacher availability ---
    ("Student support & teacher availability", "De Boer", "naam",
     "Docent De Boer is altijd bereikbaar voor vragen. Ze neemt oprecht de tijd voor elke individuele student."),
    ("Student support & teacher availability", "Larsson", "naam",
     "Teacher Larsson is very supportive and always available during office hours. A truly great mentor for students."),
    ("Student support & teacher availability", "Visser|R10|4e etage", "naam|gebouw|indirect_locatie",
     "Meneer Visser zit op de 4e etage in R10 en is altijd welkom voor een gesprek, zelfs zonder afspraak."),
    ("Student support & teacher availability", "Hendriks|c.hendriks@fontys.nl", "naam|email",
     "Mevrouw Hendriks (c.hendriks@fontys.nl) reageert snel op mail en helpt studenten goed op weg."),
    ("Student support & teacher availability", "Martens|lokaal 3.14|rolstoelgebruiker", "naam|gebouw|toegankelijkheid",
     "Docent Martens in lokaal 3.14 heeft speciaal voor mij als rolstoelgebruiker een toegankelijke zitplaats vooraan in de klas geregeld."),
    ("Student support & teacher availability", "Elbers|543210", "naam|studentnummer",
     "Student 543210 wil graag bedanken: docent Elbers heeft me enorm geholpen bij het afronden van mijn scriptie."),
    ("Student support & teacher availability", "445678|Bakker|slechthorend", "studentnummer|naam|toegankelijkheid",
     "Student 445678 heeft meerdere keren aangegeven slechthorend te zijn, maar docent Bakker weigert een microfoon te gebruiken in de grote collegezaal."),
    ("Student support & teacher availability", "Prins", "naam",
     "Docent Prins is nooit beschikbaar buiten de lesuren. Dit maakt het voor studenten met vragen erg moeilijk."),
    ("Student support & teacher availability", "Vermeer", "naam",
     "Docent Vermeer is moeilijk te bereiken. De spreekuren worden te vaak en te laat afgezegd via een kort berichtje."),
    ("Student support & teacher availability", "a.de_boer@student.fontys.nl|motorische beperking", "email|toegankelijkheid",
     "Ik heb via a.de_boer@student.fontys.nl doorgegeven dat ik een motorische beperking heb en extra tijd nodig heb bij toetsen. Er is nooit een aanpassing gedaan."),

    # --- Technology & online resources ---
    ("Technology & online resources", "Jansen", "naam",
     "De online omgeving die docent Jansen gebruikt is onduidelijk georganiseerd en moeilijk te navigeren voor studenten."),
    ("Technology & online resources", "Thompson", "naam",
     "The online resources provided by Thompson are completely outdated. We need more current and relevant materials."),
    ("Technology & online resources", "Boer", "naam",
     "Mevrouw Boer maakt goed gebruik van digitale tools. De online lessen zijn interactief en goed doordacht."),
    ("Technology & online resources", "Smit", "naam",
     "Docent Smit heeft de videoopnames van de lessen niet beschikbaar gesteld terwijl dit expliciet was beloofd."),
    ("Technology & online resources", "Willems|TQ 3.14", "naam|gebouw",
     "In TQ 3.14 is het wifi slecht. Docent Willems gebruikt online tools maar die werken tijdens de les niet."),
    ("Technology & online resources", "Prins|@r_prins_fontys", "naam|username",
     "Docent Prins (@r_prins_fontys) heeft alle cursusmateriaal achter een wachtwoord gezet zonder verdere uitleg."),
    ("Technology & online resources", "@tech_support_fontys", "username",
     "@tech_support_fontys heeft op mijn melding nooit gereageerd over de storingen in het online leerplatform."),
    ("Technology & online resources", "j.de_groot@fontys.nl", "email",
     "Ik heb j.de_groot@fontys.nl gemaild over ernstige problemen met de digitale leeromgeving maar geen reactie."),
    ("Technology & online resources", "321098", "studentnummer",
     "Student 321098: het online toetsplatform crasht constant tijdens examens, dit is echt volledig onacceptabel."),
    ("Technology & online resources", "p punt vermeer apenstaartje fontys punt nl", "obfuscated_email",
     "De docent is bereikbaar via p punt vermeer apenstaartje fontys punt nl maar dit staat nergens vermeld op Canvas."),

    # --- Diversity & inclusivity ---
    ("Diversity & inclusivity", "Peters", "naam",
     "Docent Peters maakt soms opmerkingen die niet inclusief zijn, met name tegenover internationale studenten."),
    ("Diversity & inclusivity", "Adams", "naam",
     "Mevrouw Adams is very inclusive and makes sure every student feels genuinely welcome regardless of their background."),
    ("Diversity & inclusivity", "Van Dijk", "naam",
     "Docent Van Dijk houdt rekening met verschillende culturele achtergronden. Zijn aanpak is altijd respectvol."),
    ("Diversity & inclusivity", "Williams", "naam",
     "Williams creates a very safe learning environment where every student is treated equally and with full respect."),
    ("Diversity & inclusivity", "Bakker", "naam",
     "Docent Bakker gebruikt altijd inclusief taalgebruik en stimuleert actief diversiteit in de klas."),
    ("Diversity & inclusivity", "Claes|789012", "naam|studentnummer",
     "Student 789012 ervaart dat docent Claes meer aandacht besteedt aan Nederlandstalige dan anderstalige studenten."),
    ("Diversity & inclusivity", "334455|Claes|ADHD", "studentnummer|naam|gezondheid",
     "Ik (studentnummer 334455) heb ADHD en vroeg docent Claes om extra tijd bij de toets. Dit werd geweigerd zonder enige toelichting over de procedure."),
    ("Diversity & inclusivity", "De Vries|diabetes", "naam|gezondheid",
     "Docent De Vries reageerde zichtbaar geïrriteerd toen ik tijdens de les mijn bloedsuiker moest meten. Als diabetespatiënt heb ik hier recht op."),
    ("Diversity & inclusivity", "hoofddoek|vooraan zit", "indirect_fysiek|indirect_fysiek",
     "Een specifieke studente die altijd vooraan zit en een hoofddoek draagt wordt door de docent anders behandeld."),
    ("Diversity & inclusivity", "de baard", "indirect_fysiek",
     "De mannelijke docent met de baard die Communicatie geeft negeert consistent vrouwelijke studenten tijdens discussies."),
]


def build_te_detecteren_label(te_det: str, pii_type: str) -> str:
    """Combine type:value pairs into a single string matching the example format.
    E.g. te_det='Smith|j.smith@fontys.nl', pii_type='naam|email'
    -> 'naam:Smith, email:j.smith@fontys.nl'
    """
    values = te_det.split("|")
    types = pii_type.split("|")
    return ", ".join(f"{t}:{v}" for t, v in zip(types, values))


# Build final rows: (theme, voorkomende_tekst, open_antwoord)
final_rows = [
    (theme, build_te_detecteren_label(te_det, pii_type), feedback)
    for theme, te_det, pii_type, feedback in rows
]

# --- Write CSV ---
csv_path = "C:/fontys/semester_4/group/agents/privacy_officer/test_dataset_v2.csv"
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
    writer.writerow(["thema's", "voorkomende tekst van pii of indirect wat eruit gehaald moet worden", "open antwoord"])
    for row in final_rows:
        writer.writerow(row)

# --- Write Excel ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

xlsx_path = "C:/fontys/semester_4/group/agents/privacy_officer/test_dataset_v2.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Dataset"

headers = ["thema's", "voorkomende tekst van pii of indirect wat eruit gehaald moet worden", "open antwoord"]
ws.append(headers)

# Header styling
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center")

# Theme colors for alternating rows per theme
theme_colors = {
    "Slow grading by Smith":                   "FCE4D6",
    "Insufficient contact hours":              "FFF2CC",
    "Course depth & quality (positive)":       "E2EFDA",
    "Unclear assessment criteria":             "DDEBF7",
    "Poor teacher communication":              "FCE4D6",
    "Teaching style & explanation quality":    "FFF2CC",
    "Workload & time management":              "E2EFDA",
    "Student support & teacher availability":  "DDEBF7",
    "Technology & online resources":           "FCE4D6",
    "Diversity & inclusivity":                 "FFF2CC",
}

for i, (theme, te_det_label, feedback) in enumerate(final_rows, 2):
    ws.append([theme, te_det_label, feedback])
    color = theme_colors.get(theme, "FFFFFF")
    fill = PatternFill("solid", fgColor=color)
    for cell in ws[i]:
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Column widths
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 80

wb.save(xlsx_path)

# --- Validate ---
with open(csv_path, encoding="utf-8-sig") as f:
    parsed = list(csv.reader(f, delimiter=";"))

print(f"Rijen (incl. header): {len(parsed)}")
print(f"Kolommen: {parsed[0]}")
print()

errors = []
for i, row in enumerate(parsed[1:], 2):
    theme, te_det_label, feedback = row
    pairs = [p.strip() for p in te_det_label.split(",")]
    for pair in pairs:
        if ":" not in pair:
            continue
        _, value = pair.split(":", 1)
        if value.lower() not in feedback.lower():
            errors.append(f"Rij {i} [{theme[:30]}]: '{value}' NIET gevonden in feedback")

if errors:
    print("FOUTEN:")
    for e in errors:
        print(" ", e)
else:
    print("Validatie OK: alle te_detecteren waarden staan letterlijk in open antwoord.")

types_flat = []
for row in parsed[1:]:
    pairs = [p.strip() for p in row[1].split(",")]
    for pair in pairs:
        if ":" in pair:
            t, _ = pair.split(":", 1)
            types_flat.append(t.strip())
print()
print("PII type verdeling:")
for t, c in sorted(Counter(types_flat).items(), key=lambda x: -x[1]):
    print(f"  {c:3}x  {t}")

print(f"\nBestanden geschreven:")
print(f"  CSV:  {csv_path}")
print(f"  XLSX: {xlsx_path}")
