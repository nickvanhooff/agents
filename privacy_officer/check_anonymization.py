"""
check_anonymization.py
----------------------
Loads an anonymized CSV (output of the Privacy Officer pipeline) and checks
whether each PII value listed in 'voorkomende tekst van pii of indirect...'
has actually been removed from 'anonymized_open antwoord'.

Adds three new columns:
  - geanonimiseerd     : True if ALL listed PII values are gone, False otherwise
  - gemiste_pii        : comma-separated list of values that are still present
  - extra_geanonimiseerd : comma-separated list of UNEXPECTED replacements (false positives)

Row colors in Excel:
  - Green  : fully anonymized, no unexpected extras
  - Orange : fully anonymized, but extra text was removed too
  - Red    : PII still present in the output

Exports:
  - <input>_checked.xlsx  with green/orange/red row highlighting
  - <input>_checked.csv   for programmatic use
  - Prints recall / false-positive summary to terminal
"""

import sys
import re
import difflib
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import Counter
from pathlib import Path

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
INPUT_FILE = Path("C:/fontys/semester_4/group/agents/privacy_officer/safe_test_dataset_v2 (1).csv")

COL_PII    = "voorkomende tekst van pii of indirect wat eruit gehaald moet worden"
COL_ANON   = "anonymized_open antwoord"
COL_RESULT = "geanonimiseerd"
COL_MISSED = "gemiste_pii"
COL_EXTRA  = "extra_geanonimiseerd"

TAG_PATTERN = re.compile(r'\[(?:NAME|TITLE|PII|LOCATION|COURSE[/_]?DEPT|PHYSICAL_DESCRIPTOR|STUDENT_NR|EMAIL|PHONE)[^\]]*\]')

# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------
for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
    for sep in (",", ";", "\t"):
        try:
            df = pd.read_csv(INPUT_FILE, encoding=enc, sep=sep)
            if len(df.columns) > 1:
                print(f"Loaded {len(df)} rows  (encoding={enc}, sep={sep!r})")
                break
        except Exception:
            continue
    else:
        continue
    break
else:
    print("ERROR: could not parse input file.")
    sys.exit(1)

if COL_PII not in df.columns or COL_ANON not in df.columns:
    print(f"ERROR: expected columns not found.\nFound: {list(df.columns)}")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Check per row
# ---------------------------------------------------------------------------
def check_row(pii_label: str, anonymized: str):
    """
    Returns (all_removed: bool, missed: list[str])
    pii_label format: "naam:Smith, email:j.smith@fontys.nl, indirect_fysiek:rode das"
    """
    if not isinstance(anonymized, str):
        anonymized = ""

    missed = []
    pairs = [p.strip() for p in str(pii_label).split(",")]
    for pair in pairs:
        if ":" not in pair:
            continue
        _, value = pair.split(":", 1)
        value = value.strip()
        if not value:
            continue
        # Case-insensitive check: value should NOT appear in anonymized text
        if value.lower() in anonymized.lower():
            missed.append(value)

    return (len(missed) == 0), missed


def find_extra_replacements(original: str, anonymized: str, expected_values: list) -> list:
    """
    Returns a list of words that were unexpectedly replaced by a [TAG].
    Uses word-level difflib to find replaced segments, then checks each word
    individually against the expected PII values so that mixed segments like
    'docent Jansen' → '[TITLE] [NAME]' correctly flag only 'docent'.
    """
    if not isinstance(original, str) or not isinstance(anonymized, str):
        return []
    if not original.strip() or anonymized.startswith('[NEEDS_REVIEW'):
        return []

    expected_lower = [v.lower() for v in expected_values if v]
    orig_words = original.split()
    anon_words = anonymized.split()

    extra = []
    matcher = difflib.SequenceMatcher(None, orig_words, anon_words, autojunk=False)

    for opcode, i1, i2, j1, j2 in matcher.get_opcodes():
        if opcode != 'replace':
            continue
        anon_segment = ' '.join(anon_words[j1:j2])
        # Only care about replacements that produced a [TAG]
        if not TAG_PATTERN.search(anon_segment):
            continue
        # Check each word individually — avoids masking 'docent' because 'Jansen' is expected
        for word in orig_words[i1:i2]:
            clean = word.lower().strip('.,!?;:()"\'')
            if len(clean) <= 2:
                continue
            was_expected = any(
                exp in clean or clean in exp
                for exp in expected_lower
            )
            if not was_expected:
                extra.append(word)

    return extra


results = []
missed_all = []
extras_all = []

for _, row in df.iterrows():
    ok, missed = check_row(row[COL_PII], row[COL_ANON])
    results.append(ok)
    missed_all.append(", ".join(missed) if missed else "")

    # Parse expected values from PII label
    expected = []
    for pair in str(row[COL_PII]).split(","):
        pair = pair.strip()
        if ":" in pair:
            _, val = pair.split(":", 1)
            val = val.strip()
            if val:
                expected.append(val)

    extras = find_extra_replacements(str(row.get("open antwoord", "")), str(row[COL_ANON]), expected)
    extras_all.append(", ".join(extras) if extras else "")

df[COL_RESULT] = results
df[COL_MISSED] = missed_all
df[COL_EXTRA]  = extras_all

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
total         = len(df)
fully_anon    = df[COL_RESULT].sum()
not_anon      = total - fully_anon
recall        = fully_anon / total * 100

has_extras    = df[COL_EXTRA].astype(str).str.len() > 0
extra_count   = int(has_extras.sum())

print()
print("=" * 55)
print(f"  Total rows          : {total}")
print(f"  Fully anonymized    : {fully_anon}  ({recall:.1f}%  recall)")
print(f"  PII still present   : {not_anon}")
print(f"  Rows with extras    : {extra_count}  (false positives)")
print("=" * 55)

# Per-theme breakdown
print()
print("Per thema:")
theme_col = df.columns[0]
for theme, grp in df.groupby(theme_col):
    ok = grp[COL_RESULT].sum()
    extras_in_theme = int((grp[COL_EXTRA].astype(str).str.len() > 0).sum())
    extras_note = f"  (+{extras_in_theme} extra)" if extras_in_theme else ""
    print(f"  {ok:2}/{len(grp):2}  {theme}{extras_note}")

# Most-missed PII types
all_missed_values = [v for row in missed_all for v in row.split(", ") if v]
if all_missed_values:
    print()
    print("Meest gemiste PII-waarden:")
    for val, cnt in Counter(all_missed_values).most_common(10):
        print(f"  {cnt}x  '{val}'")

# Most over-anonymized (false positives)
all_extras = [v for row in extras_all for v in row.split(", ") if v]
if all_extras:
    print()
    print("Meest onverwacht verwijderde woorden (false positives):")
    for val, cnt in Counter(all_extras).most_common(10):
        print(f"  {cnt}x  '{val}'")

# ---------------------------------------------------------------------------
# Export CSV
# ---------------------------------------------------------------------------
out_csv = INPUT_FILE.with_name(INPUT_FILE.stem + "_checked.csv")
df.to_csv(out_csv, index=False, encoding="utf-8-sig")
print(f"\nCSV opgeslagen: {out_csv}")

# ---------------------------------------------------------------------------
# Export Excel with color coding
# ---------------------------------------------------------------------------
out_xlsx = INPUT_FILE.with_name(INPUT_FILE.stem + "_checked.xlsx")

GREEN  = "C6EFCE"   # light green  — fully anonymized, no extras
ORANGE = "FFEB9C"   # light orange — fully anonymized, but unexpected text removed too
RED    = "FFC7CE"   # light red    — PII still present
HEADER_COLOR = "1F3864"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Anonymization Check"

# Header
ws.append(list(df.columns))
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(wrap_text=True, vertical="center")

# Rows
for _, row in df.iterrows():
    ws.append(list(row))

# Color last two data columns + entire row tint
result_col_idx = df.columns.get_loc(COL_RESULT) + 1  # 1-based

for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
    is_ok     = bool(row[COL_RESULT])
    has_extra = bool(str(row[COL_EXTRA]).strip())

    if not is_ok:
        row_color  = RED
        font_color = "9C0006"
    elif has_extra:
        row_color  = ORANGE
        font_color = "7D4E00"
    else:
        row_color  = GREEN
        font_color = "375623"

    # Tint entire row
    for cell in ws[row_idx]:
        cell.fill = PatternFill("solid", fgColor=row_color)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Bold the result cell
    ws.cell(row=row_idx, column=result_col_idx).font = Font(
        bold=True,
        color=font_color
    )

# Column widths
col_widths = {1: 30, 2: 50, 3: 55, 4: 55, 5: 14, 6: 35}
for col, width in col_widths.items():
    if col <= ws.max_column:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Freeze header
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet("Summary")
ws2.append(["Metric", "Value"])
ws2.append(["Total rows", total])
ws2.append(["Fully anonymized", int(fully_anon)])
ws2.append(["PII still present", int(not_anon)])
ws2.append(["Recall (%)", round(recall, 1)])
ws2.append(["Rows with false positives", extra_count])
ws2.append([])
ws2.append(["Theme", "Anonymized", "Total", "Recall (%)", "False Positives"])
for theme, grp in df.groupby(theme_col):
    ok  = int(grp[COL_RESULT].sum())
    n   = len(grp)
    fp  = int((grp[COL_EXTRA].astype(str).str.len() > 0).sum())
    ws2.append([theme, ok, n, round(ok / n * 100, 1), fp])

if all_extras:
    ws2.append([])
    ws2.append(["Top onverwachte vervangingen", "Aantal"])
    for val, cnt in Counter(all_extras).most_common(10):
        ws2.append([val, cnt])

for cell in ws2[1]:
    cell.font = Font(bold=True)
for cell in ws2["A"]:
    if cell.row > 1:
        cell.font = Font(bold=True)

wb.save(out_xlsx)
print(f"Excel opgeslagen: {out_xlsx}")
